"""Write one dataset profile with one engine, in its own process, and report
wall time + peak RSS. One engine per process so peak RSS is attributable.

    python bench_engines.py <engine> <profile>

engines : fastxlsx | openpyxl_writeonly | xlsxwriter | xlsxwriter_constant | pandas
profiles: mixed | numeric | strings | wide

Output (parseable): RESULT <engine> <profile> <time_s> <peak_rss_mb>
"""

import resource
import sys
import time

# Each profile holds ~1,000,000 cells so engines are compared on equal work.
PROFILES = {
    "mixed": (200_000, 5, lambda i: [f"patient_{i}", i, i * 1.5, i % 2 == 0, "NSQIP"]),
    "numeric": (200_000, 5, lambda i: [i, i * 1.5, i * 2, i * 0.001, float(i % 7)]),
    "strings": (200_000, 5, lambda i: [f"a{i}", f"b{i}", f"c{i}", f"d{i}", f"e{i}"]),
    "wide": (20_000, 50, lambda i: [_wide_cell(i, j) for j in range(50)]),
}


def _wide_cell(row, col):
    kind = col % 4
    if kind == 0:
        return row + col
    if kind == 1:
        return f"r{row}c{col}"
    if kind == 2:
        return (row + col) * 0.5
    return (row + col) % 2 == 0


def run_fastxlsx(rows, make_row, path):
    import fastxlsx

    wb = fastxlsx.Workbook()
    ws = wb.add_sheet("Sheet1")
    for i in range(rows):
        ws.append(make_row(i))
    wb.save(path)


def run_openpyxl_writeonly(rows, make_row, path):
    from openpyxl import Workbook

    wb = Workbook(write_only=True)
    ws = wb.create_sheet("Sheet1")
    for i in range(rows):
        ws.append(make_row(i))
    wb.save(path)


def run_xlsxwriter(rows, make_row, path, constant_memory=False):
    import xlsxwriter

    wb = xlsxwriter.Workbook(path, {"constant_memory": constant_memory})
    ws = wb.add_worksheet("Sheet1")
    for i in range(rows):
        ws.write_row(i, 0, make_row(i))
    wb.close()


def run_pandas(rows, make_row, path):
    import pandas as pd

    df = pd.DataFrame([make_row(i) for i in range(rows)])
    df.to_excel(path, index=False, header=False)


ENGINES = {
    "fastxlsx": run_fastxlsx,
    "openpyxl_writeonly": run_openpyxl_writeonly,
    "xlsxwriter": lambda r, m, p: run_xlsxwriter(r, m, p, constant_memory=False),
    "xlsxwriter_constant": lambda r, m, p: run_xlsxwriter(r, m, p, constant_memory=True),
    "pandas": run_pandas,
}


def peak_rss_mb():
    rss = resource.getrusage(resource.RUSAGE_SELF).ru_maxrss
    # macOS reports bytes, Linux reports kilobytes.
    return rss / 1_000_000 if sys.platform == "darwin" else rss / 1_000


def main():
    engine, profile = sys.argv[1], sys.argv[2]
    rows, _cols, make_row = PROFILES[profile]
    path = f"/tmp/bench_{engine}_{profile}.xlsx"

    start = time.perf_counter()
    ENGINES[engine](rows, make_row, path)
    elapsed = time.perf_counter() - start

    print(f"RESULT {engine} {profile} {elapsed:.3f} {peak_rss_mb():.1f}")


if __name__ == "__main__":
    main()
